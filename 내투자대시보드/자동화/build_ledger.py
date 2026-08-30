#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""연별 원장 조립기 (자동화 파이프라인) — 간소화 버전.

`데이터/`에 누적된 일별 공개 데이터에서 **연말 값**을 뽑아, 대시보드가 읽는 원장 양식의
자동 채움 시트를 만들어 `자동수집_원장.xlsx`로 저장합니다.

자동으로 채우는 것 (공개 데이터):
  · 연말시장  : 연말 USD/KRW · S&P500 · NASDAQ · 다우
  · 연말보유  : tickers.txt의 (티커, 수량) × 연말 종가 → 평가액USD

당신이 채우는 것 (개인 데이터, 연 1회):
  · 연별스냅샷: 자산군별 연말평가액 + 당해순입금 (+ MMF·RP·현금 잔고)
  · 목표      : 목표금액·목표일·월저축액

원장에 **없는 것**: 시장지표(미국채10년·금·WTI)·Fear&Greed·미국 기준금리.
이것들은 `build_dashboard.py` 가 대시보드 HTML 안에 매일 직접 심습니다.
"""
from __future__ import annotations

import csv
import os
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "데이터")
OUT = os.path.join(HERE, "자동수집_원장.xlsx")
TICKERS = os.path.join(HERE, "tickers.txt")

HFILL = PatternFill("solid", fgColor="1F2A44")
HFONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="center")


def read_csv(name):
    p = os.path.join(DATA_DIR, name)
    if not os.path.exists(p):
        return []
    with open(p, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def year_end_map(rows, valcols):
    """{YYYY: {col: value}} — 각 연도의 마지막 날짜 값."""
    best = {}
    for r in rows:
        d = r.get("date", "")
        if len(d) < 4:
            continue
        y = d[:4]
        if y not in best or d > best[y]["date"]:
            best[y] = r
    return {y: {c: r.get(c) for c in valcols} for y, r in best.items()}


def read_tickers():
    out = []
    if os.path.exists(TICKERS):
        for line in open(TICKERS, encoding="utf-8"):
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            p = line.split()
            out.append((p[0], float(p[1]) if len(p) > 1 else 0.0, p[2] if len(p) > 2 else ""))
    return out


def header(ws, title, sub, cols, widths):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = sub
    ws["A2"].font = Font(italic=True, color="666666")
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=4, column=j, value=c)
        cell.fill = HFILL
        cell.font = HFONT
        cell.alignment = WRAP
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    fx = year_end_map(read_csv("환율_일별.csv"), ["usdkrw"])
    bm = year_end_map(read_csv("벤치마크_일별.csv"), ["sp", "nasdaq", "dow"])
    px_rows = read_csv("시세_일별.csv")
    by_ticker = defaultdict(list)
    for r in px_rows:
        by_ticker[r.get("ticker", "")].append(r)
    px_ye = {t: year_end_map(rows, ["close"]) for t, rows in by_ticker.items()}
    tickers = read_tickers()

    wb = openpyxl.Workbook()
    g = wb.active
    g.title = "안내"
    g["A1"] = "자동수집 원장 (공개 데이터 자동 채움) · 연별"
    g["A1"].font = Font(bold=True, size=13)
    for i, line in enumerate([
        "'연별스냅샷'과 '목표'만 채우면 대시보드에 넣을 수 있습니다.",
        "",
        "연말시장·연말보유는 자동으로 채워집니다.",
        "시장지표·Fear&Greed·미국 기준금리는 원장에 없습니다 — 대시보드 HTML에 매일 자동으로 심깁니다.",
    ], start=2):
        c = g.cell(row=i, column=1, value=line)
        if i == 2:
            c.font = Font(italic=True, color="666666")
    g.column_dimensions["A"].width = 78

    # 연별스냅샷 — 헤더만 (사용자 입력, 연 1회)
    ws = wb.create_sheet("연별스냅샷")
    header(ws, "연별 스냅샷 — 자산군별로 연 1줄 (직접 입력)",
           "기준연(YYYY)·자산군·통화·연말평가액·당해순입금. 자산군 6종만.",
           ["기준연\n(YYYY)", "자산군", "통화", "연말평가액", "당해순입금", "계좌(선택)", "비고"],
           [12, 12, 8, 15, 15, 12, 24])

    # 연말시장 (자동) — 환율 + 지수를 한 장에
    years = sorted(set(fx) | set(bm))
    mk = wb.create_sheet("연말시장")
    header(mk, "연말 시장 기준값 (자동수집)", "연 1줄. 개인 성과를 환산·비교하는 데만 씁니다.",
           ["기준연\n(YYYY)", "USD/KRW\n(연말)", "S&P 500", "NASDAQ\n종합", "다우존스\n산업평균"],
           [12, 14, 12, 12, 14])
    for i, y in enumerate(years):
        mk.cell(row=5 + i, column=1, value=y)
        mk.cell(row=5 + i, column=2, value=_num((fx.get(y) or {}).get("usdkrw")))
        mk.cell(row=5 + i, column=3, value=_num((bm.get(y) or {}).get("sp")))
        mk.cell(row=5 + i, column=4, value=_num((bm.get(y) or {}).get("nasdaq")))
        mk.cell(row=5 + i, column=5, value=_num((bm.get(y) or {}).get("dow")))

    # 연말보유 (자동: 티커×수량×연말가)
    hs = wb.create_sheet("연말보유")
    header(hs, "연말 종목별 보유 (자동수집)", "tickers.txt의 수량 × 연말 종가.",
           ["기준연\n(YYYY)", "티커", "종목명", "수량", "종가\n(USD)", "평가액\n(USD)", "자산군"],
           [12, 10, 22, 10, 12, 14, 12])
    r = 5
    all_years = sorted(set().union(*[set(px_ye.get(t, {})) for t, _, _ in tickers]) if tickers else set())
    for y in all_years:
        for tk, qty, cls in tickers:
            m = px_ye.get(tk, {}).get(y)
            if not m:
                continue
            px = _num(m["close"])
            if px is None:
                continue
            hs.cell(row=r, column=1, value=y)
            hs.cell(row=r, column=2, value=tk)
            hs.cell(row=r, column=3, value="")
            hs.cell(row=r, column=4, value=qty)
            hs.cell(row=r, column=5, value=round(px, 2))
            hs.cell(row=r, column=6, value=round(px * qty, 2))
            hs.cell(row=r, column=7, value=cls)
            r += 1

    # 배당 — 헤더만 (선택 입력)
    dv = wb.create_sheet("배당")
    header(dv, "배당 수령 (연 단위 · 선택)", "연도·티커·그해 받은 배당금 합계(USD).",
           ["기준연\n(YYYY)", "티커", "배당금\n(USD)", "비고"], [12, 10, 14, 18])

    # 목표 — 헤더만
    gs = wb.create_sheet("목표")
    header(gs, "재무 목표 (직접 입력)", "목표금액·목표일·월저축액·가정수익률.",
           ["목표명", "목표금액\n(KRW)", "목표일\n(YYYY-MM)", "월 저축액\n(KRW)", "가정 연복리\n수익률", "비고"],
           [16, 16, 12, 14, 14, 20])

    wb.save(OUT)
    print(f"생성: {OUT}  (연말시장 {len(years)}년 · 종목 {len(tickers)}개 · 시트 {len(wb.sheetnames)}장)")


if __name__ == "__main__":
    main()
