#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""샘플 투자원장 생성기 (테스트/데모용) — 연별·간소화 버전.

실제 원장 없이 대시보드를 테스트할 수 있도록 **더미 데이터**를 만들어
`투자원장_샘플.xlsx` 로 저장합니다.

## 원장이 얇아진 이유

시장 데이터(시장지표·Fear&Greed·기준금리)는 이제 **대시보드 HTML 안에 매일 자동으로
심깁니다**. 그래서 원장에서 통째로 빠졌습니다. 남은 건 "시장이 대신 알 수 없는 것"뿐입니다.

    연별스냅샷   자산군별 연말평가액 + 그해 순입금   ← 오직 이것만 직접 입력
    연말시장     연말 USD/KRW·S&P·NASDAQ·다우      ← 자동수집이 채움 (연 1줄)
    연말보유     티커·수량·종가·평가액              ← 자동수집이 채움 (선택)
    거래일지     거래일·구분·티커·수량·단가          ← 직접 입력 (선택 · 매매 판정용)
    배당         연도·티커·배당금                   ← 선택
    목표         목표금액·목표일·월저축액           ← 한 번만

시트 9장·약 800줄 → **7장·약 150줄**. 헤더는 4행, 데이터는 5행부터(§2).
seed 고정이라 실행할 때마다 동일하게 재현되고, 전부 **가짜 값**이라 커밋해도 안전합니다.

사용법:
    python3 샘플원장생성.py               # 투자원장_샘플.xlsx (7년, 2019~2025)
    python3 샘플원장생성.py 8 파일.xlsx    # 연수/파일명 지정
"""
from __future__ import annotations

import random
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SEED = 20260725
random.seed(SEED)

LAST_YEAR = 2025

# 자산군 6종 (§2). (자산군, 통화, 시드 연말평가액, 연평균수익률, 연변동성, 연 순입금)
ASSETS = [
    ("해외주식",   "USD",     14000,   0.115, 0.170,   6000),
    ("해외ETF",   "USD",      9000,   0.100, 0.150,   4200),
    ("미국장기채", "USD",      8000,   0.015, 0.110,   2400),
    ("MMF",       "KRW",   7_000_000, 0.033, 0.006, 4_000_000),
    ("RP",        "KRW",   5_000_000, 0.030, 0.005,       0),
    ("달러현금",   "USD",      2500,   0.000, 0.003,       0),
]

# (티커, 종목명, 시작수량, 시작가, 연평균수익률, 연변동성, 자산군, 배당수익률)
HOLDINGS = [
    ("VOO", "Vanguard S&P 500 ETF", 20, 250, 0.125, 0.165, "해외ETF", 0.013),
    ("QQQ", "Invesco QQQ Trust", 10, 170, 0.150, 0.210, "해외ETF", 0.006),
    ("AAPL", "Apple Inc.", 40, 55, 0.140, 0.230, "해외주식", 0.005),
    ("MSFT", "Microsoft Corp.", 18, 110, 0.145, 0.200, "해외주식", 0.008),
    ("TLT", "iShares 20+ Yr Treasury", 90, 130, -0.020, 0.130, "미국장기채", 0.038),
]

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="center")


def year_list(n: int, last: int = LAST_YEAR) -> list[int]:
    return list(range(last - n + 1, last + 1))


def write_header(ws, title, subtitle, headers, widths):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="666666")
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


def build(years: list[int]):
    wb = openpyxl.Workbook()

    # ── 안내 ────────────────────────────────────────────────────────────
    g = wb.active
    g.title = "안내"
    g["A1"] = "투자 대시보드 원장 — 샘플(더미) 데이터 · 연별 간소화"
    g["A1"].font = Font(bold=True, size=13)
    g["A2"] = "직접 채우는 건 '연별스냅샷'과 '목표' 둘뿐입니다. 나머지는 자동수집이 채웁니다."
    g["A2"].font = Font(italic=True, color="666666")
    guide = [
        "",
        "[직접 입력]",
        "  연별스냅샷 — 자산군별 연말평가액 + 그해 순입금 (연 1회)",
        "  목표       — 목표금액·목표일·월저축액 (한 번만)",
        "",
        "  거래일지   — 사고판 기록 (있으면 '매매일지' 탭이 켜집니다 · 없어도 무방)",
        "",
        "[자동수집이 채움]",
        "  연말시장   — 연말 USD/KRW·S&P500·NASDAQ·다우",
        "  연말보유   — 티커·수량·종가·평가액",
        "  배당       — 연도·티커·배당금",
        "",
        "[원장에 없음 — 대시보드 HTML에 매일 자동으로 심깁니다]",
        "  시장지표(미국채10년·금·WTI) · Fear&Greed · 미국 기준금리",
        "",
        "자산군은 아래 6종만 사용합니다:",
    ]
    for i, line in enumerate(guide, start=4):
        g.cell(row=i, column=1, value=line)
    for i, (name, *_) in enumerate(ASSETS, start=4 + len(guide)):
        g.cell(row=i, column=1, value="  · " + name)
    g.column_dimensions["A"].width = 64

    # ── 연별스냅샷 (개인 성과 · 직접 입력) ──────────────────────────────
    ws = wb.create_sheet("연별스냅샷")
    write_header(
        ws, "연별 스냅샷 — 자산군별로 연 1줄 (직접 입력)",
        "기준연은 YYYY. 금액은 해당 통화 기준 그대로. 연말평가액 + 그해 순입금.",
        ["기준연\n(YYYY)", "자산군", "통화", "연말평가액", "당해순입금", "계좌(선택)", "비고"],
        [12, 12, 8, 15, 15, 12, 26],
    )
    prev_ev = {a[0]: 0.0 for a in ASSETS}
    r = 5
    for yi, year in enumerate(years):
        for (name, cur, seed_amt, mu, sig, dep) in ASSETS:
            if yi == 0:
                ev = float(seed_amt)
                cf = ev  # 첫 해: 순입금 = 평가액
            else:
                ret = random.gauss(mu, sig)
                if year == 2022 and name in ("해외주식", "해외ETF", "미국장기채"):
                    ret = -abs(random.gauss(0.17, 0.05))   # 2022 위험자산 하락(낙폭 데모)
                cf = float(dep)
                if random.random() < 0.25:            # 가끔 목돈 입금/인출
                    cf += random.choice([-1, 1, 1]) * seed_amt * random.uniform(0.10, 0.40)
                ev = max(prev_ev[name] * (1 + ret) + cf, 0.0)
            prev_ev[name] = ev
            acct = "미래에셋" if cur == "USD" else "NH투자"
            ws.cell(row=r, column=1, value=str(year))   # 텍스트 "YYYY"
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=cur)
            ws.cell(row=r, column=4, value=round(ev, 2) if cur == "USD" else round(ev))
            ws.cell(row=r, column=5, value=round(cf, 2) if cur == "USD" else round(cf))
            ws.cell(row=r, column=6, value=acct)
            ws.cell(row=r, column=7, value="첫 해라 순입금=평가액" if yi == 0 else "")
            r += 1

    # ── 연말시장 (환율+지수 한 장 · 자동수집) ───────────────────────────
    mk = wb.create_sheet("연말시장")
    write_header(
        mk, "연말 시장 기준값 (자동수집)", "연 1줄. 개인 성과를 환산·비교하는 데만 씁니다.",
        ["기준연\n(YYYY)", "USD/KRW\n(연말)", "S&P 500", "NASDAQ\n종합", "다우존스\n산업평균"],
        [12, 14, 12, 12, 14],
    )
    fx, sp, nq, dj = 1160.0, 3230.0, 8970.0, 28540.0
    for i, year in enumerate(years):
        if i > 0:
            fx = min(1480.0, max(1050.0, fx * (1 + random.gauss(0.035, 0.07))))
            sp *= 1 + random.gauss(0.115, 0.170)
            nq *= 1 + random.gauss(0.140, 0.220)
            dj *= 1 + random.gauss(0.090, 0.150)
        mk.cell(row=5 + i, column=1, value=str(year))
        mk.cell(row=5 + i, column=2, value=round(fx))
        mk.cell(row=5 + i, column=3, value=round(sp, 2))
        mk.cell(row=5 + i, column=4, value=round(nq, 2))
        mk.cell(row=5 + i, column=5, value=round(dj, 2))

    # ── 월별 시세 경로 (거래일지용) ────────────────────────────────────
    # 매매 판정을 하려면 '거래한 달의 가격'이 필요합니다. 월 단위로 시뮬레이션한 뒤
    # 12월 값을 그대로 연말 종가로 씁니다 — 연말보유와 거래일지가 어긋나지 않게.
    px_path: dict[str, dict[str, float]] = {}     # {티커: {"YYYY-MM": 가격}}
    for (tkr, nm, qty0, px0, mu, sig, cls, dy) in HOLDINGS:
        mu_m, sig_m = mu / 12.0, sig / (12 ** 0.5)
        px = float(px0)
        path = {}
        for yi, year in enumerate(years):
            for m in range(1, 13):
                if not (yi == 0 and m == 12):     # 첫 해 12월은 시드가를 그대로
                    r = random.gauss(mu_m, sig_m)
                    if year == 2022:              # 2022 하락장
                        r -= 0.020
                    px = max(px * (1 + r), 1.0)
                path[f"{year}-{m:02d}"] = round(px, 2)
        px_path[tkr] = path

    # ── 연말보유 (자동수집 · 선택) ──────────────────────────────────────
    hold = wb.create_sheet("연말보유")
    write_header(
        hold, "연말 종목별 보유 (자동수집 · 선택)", "종목 비중 도넛·가격수익률용. 연 1줄씩.",
        ["기준연\n(YYYY)", "티커", "종목명", "수량", "종가\n(USD)", "평가액\n(USD)", "자산군"],
        [12, 10, 26, 10, 12, 14, 12],
    )
    dv_rows = []          # 배당 시트에서 재사용
    qty_by_year: dict[str, dict[int, int]] = {}
    hr = 5
    for (tkr, nm, qty0, px0, mu, sig, cls, dy) in HOLDINGS:
        qty = qty0
        qty_by_year[tkr] = {}
        for yi, year in enumerate(years):
            if yi > 0:
                qty += random.randint(0, 6)      # 연중 추가 매수
            qty_by_year[tkr][year] = qty
            px_r = px_path[tkr][f"{year}-12"]
            hold.cell(row=hr, column=1, value=str(year))
            hold.cell(row=hr, column=2, value=tkr)
            hold.cell(row=hr, column=3, value=nm)
            hold.cell(row=hr, column=4, value=qty)
            hold.cell(row=hr, column=5, value=px_r)
            hold.cell(row=hr, column=6, value=round(px_r * qty, 2))
            hold.cell(row=hr, column=7, value=cls)
            hr += 1
            if dy > 0:
                dv_rows.append((year, tkr, round(px_r * qty * dy, 2)))

    # ── 거래일지 (직접 입력 · 선택) ─────────────────────────────────────
    # 이 시트가 있으면 '매매일지' 탭이 켜집니다. 없어도 나머지는 전부 정상 동작합니다.
    # 샘플은 잘한 결정과 아쉬운 결정을 일부러 섞습니다(둘 다 보여야 화면이 정직해집니다).
    tr = wb.create_sheet("거래일지")
    write_header(
        tr, "거래일지 — 사고판 기록 (직접 입력 · 선택)",
        "이 기록이 있으면 '그때 안 했다면?' 을 계산해 매매 판정을 보여줍니다. 단가는 수수료·세금 포함가로 적으면 더 정확합니다.",
        ["거래일\n(YYYY-MM)", "구분\n(매수/매도)", "티커", "종목명", "수량", "단가\n(USD)", "통화", "메모"],
        [13, 12, 10, 24, 10, 12, 8, 30],
    )
    NAME = {h[0]: h[1] for h in HOLDINGS}
    trades = []
    last_ym = f"{years[-1]}-12"
    for (tkr, nm, qty0, px0, mu, sig, cls, dy) in HOLDINGS:
        path = px_path[tkr]
        ref = path[last_ym]
        cand = [ym for ym in path if ym < last_ym and ym >= f"{years[0]}-04"]
        for ym in random.sample(cand, k=min(6, len(cand))):
            px = path[ym]
            side = "매도" if random.random() < 0.42 else "매수"
            qty = random.choice([3, 5, 8, 10, 12, 15, 20])
            memo = ""
            if side == "매도":
                memo = "고점 같아서 정리" if px > ref else "겁나서 정리"
            else:
                memo = "떨어진 김에 추가" if px < ref else "쫓아 들어감"
            trades.append((ym, side, tkr, NAME[tkr], qty, px, "USD", memo))
    trades.sort()
    for i, t in enumerate(trades):
        for j, v in enumerate(t, start=1):
            tr.cell(row=5 + i, column=j, value=v)

    # ── 배당 (연 단위 · 선택) ───────────────────────────────────────────
    dv = wb.create_sheet("배당")
    write_header(
        dv, "배당 수령 (연 단위 · 선택)", "연도·티커·그해 받은 배당금 합계(USD).",
        ["기준연\n(YYYY)", "티커", "배당금\n(USD)", "비고"], [12, 10, 14, 18],
    )
    for i, (year, tkr, amt) in enumerate(sorted(dv_rows)):
        dv.cell(row=5 + i, column=1, value=str(year))
        dv.cell(row=5 + i, column=2, value=tkr)
        dv.cell(row=5 + i, column=3, value=amt)

    # ── 목표 ────────────────────────────────────────────────────────────
    goal = wb.create_sheet("목표")
    write_header(
        goal, "재무 목표 · 복리 시뮬레이션 가정", "가정수익률은 대시보드 슬라이더 시작값.",
        ["목표명", "목표금액\n(KRW)", "목표일\n(YYYY-MM)", "월 저축액\n(KRW)", "가정 연복리\n수익률", "비고"],
        [16, 16, 12, 14, 14, 20],
    )
    goal.cell(row=5, column=1, value="1조 만들기")
    goal.cell(row=5, column=2, value=1_000_000_000_000)   # 기본 목표: 1조원
    goal.cell(row=5, column=3, value="2055-12")
    goal.cell(row=5, column=4, value=2_000_000)
    goal.cell(row=5, column=5, value=0.10)
    goal.cell(row=5, column=6, value="기본 목표 (샘플)")

    return wb


def main():
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 7
    out = sys.argv[2] if len(sys.argv) > 2 else "투자원장_샘플.xlsx"
    years = year_list(n)
    wb = build(years)
    wb.save(out)
    rows = n * len(ASSETS) + n + n * len(HOLDINGS) + n * sum(1 for h in HOLDINGS if h[7] > 0) + 1
    print(f"생성 완료: {out}  ({years[0]}~{years[-1]}, {n}년 · 시트 {len(wb.sheetnames)}장 · 데이터 약 {rows}줄)")


if __name__ == "__main__":
    main()
